"""Etapa 6 - Excel output.

Joins the original catalog with the Etapa 5 results (`seleccionados`) into one
DataFrame, downloads the winning image for each row to disk (needed to embed
it - Gemini's fetches in select.py are in-memory only and never saved), and
writes an .xlsx with an extra `miniatura` column holding the actual thumbnail
image via openpyxl, so a reviewer can eyeball the whole catalog without
opening a single URL.
"""
import re
from io import BytesIO
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from pipeline.select import FETCH_HEADERS

IMG_DIR = Path(__file__).resolve().parent.parent / "output" / "images"
THUMB_SIDE = 120  # px, embedded thumbnail size in the sheet


# --- Selected image download (disk, for embedding) --------------------------

def _safe(ref: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", str(ref).strip()) or "unknown"


def _image_path(ref: str) -> Path:
    return IMG_DIR / f"{_safe(ref)}.jpg"


def download_selected_image(ref: str, url: str, timeout: int = 15,
                        use_cache: bool = True, fallback_url: str = "") -> Path | None:
    """Save the winning image to disk (output/images/{ref}.jpg).

    Tries `url` (the original) and then `fallback_url` (Google's cached
    thumbnail, low-res but never hotlink-blocked). Returns the path, or None
    if both fail - the row then keeps `imagen_url` but `imagen_local`/
    `miniatura` stay empty for that product.
    """
    path = _image_path(ref)
    if use_cache and path.exists():
        return path
    img = None
    for u in (url, fallback_url):
        if not u:
            continue
        try:
            resp = requests.get(u, timeout=timeout, headers=FETCH_HEADERS)
            resp.raise_for_status()
            img = PILImage.open(BytesIO(resp.content)).convert("RGB")
            break
        except Exception:
            continue
    if img is None:
        return None

    path.parent.mkdir(parents=True, exist_ok=True)
    img.save(path, format="JPEG", quality=90)
    return path


# --- Join original catalog + Etapa 5 results ---------------------------------

def build_output_df(df: pd.DataFrame, seleccionados: dict[str, dict],
                col_ref: str = "Ref Proveedor", download_images: bool = True,
                use_cache: bool = True) -> pd.DataFrame:
    """Return a copy of `df` with the Etapa 6 columns from plan.md appended.

    `df` needs `categoria` / `termino_busqueda` / `cse_profile` already
    present (Etapas 1-2 output). `seleccionados` is the {ref: select_result}
    dict from Etapa 5 (`select.select_df`); rows with no entry there, or with
    `seleccion=None`, get empty image columns and whatever flags Etapa 5
    already attached (`sin_imagen`, `necesita_fallback`).
    """
    rows = []
    for _, row in df.iterrows():
        ref = row[col_ref]
        r = seleccionados.get(ref, {})
        sel = r.get("seleccion")
        url = sel.get("link", "") if sel else ""

        local = None
        if url and download_images:
            local = download_selected_image(
                ref, url, use_cache=use_cache,
                fallback_url=sel.get("thumbnailLink", "") if sel else "",
            )

        rows.append({
            "imagen_url": url,
            "imagen_local": str(local) if local else "",
            "categoria": row.get("categoria", ""),
            "termino_busqueda": row.get("termino_busqueda", ""),
            "perfil_cse_usado": row.get("cse_profile", ""),
            "confianza": r.get("confianza", ""),
            "flags": ", ".join(r.get("flags", [])),
            "razon_gemini": r.get("comentario", ""),
        })

    return pd.concat([df.reset_index(drop=True), pd.DataFrame(rows)], axis=1)


# --- Write .xlsx + embed thumbnails ------------------------------------------

def write_excel(out_df: pd.DataFrame, path: Path, col_imagen_local: str = "imagen_local",
            embed_images: bool = True, thumb_side: int = THUMB_SIDE) -> Path:
    """Write `out_df` to `path`, then (if `embed_images`) reopen it with
    openpyxl and drop a real thumbnail into a new `miniatura` column for every
    row whose `imagen_local` points at a file on disk.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    out_df.to_excel(path, index=False)

    if not embed_images:
        return path

    wb = load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col_idx = headers.index(col_imagen_local) + 1
    thumb_col = len(headers) + 1
    thumb_col_letter = get_column_letter(thumb_col)
    ws.cell(row=1, column=thumb_col, value="miniatura")
    ws.column_dimensions[thumb_col_letter].width = thumb_side / 7

    n_embedded = 0
    for row_i in range(2, ws.max_row + 1):
        local = ws.cell(row=row_i, column=col_idx).value
        if not local or not Path(local).exists():
            continue
        try:
            img = PILImage.open(local)
            img.thumbnail((thumb_side, thumb_side))
        except Exception:
            continue

        buf = BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        ws.add_image(XLImage(buf), f"{thumb_col_letter}{row_i}")
        ws.row_dimensions[row_i].height = thumb_side * 0.75
        n_embedded += 1

    wb.save(path)
    print(f"excel: {path} ({ws.max_row - 1} filas, {n_embedded} miniaturas incrustadas)")
    return path
